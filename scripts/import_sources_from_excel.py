import argparse
from pathlib import Path
import json

import openpyxl
from inip.db import get_conn

def norm(s):
    return (s or "").strip()

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to master_rss_list.xlsm/.xlsx")
    ap.add_argument("--sheet", default="master_rss_list", help="Worksheet name")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[args.sheet]

    headers = [norm(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}

    required = ["Feed URL", "Feed Type", "Feed Name", "Organisation"]
    for r in required:
        if r not in idx:
            raise SystemExit(f"Missing required column '{r}'. Found: {headers}")

    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in values):
            continue
        rows.append(values)

    print(f"Rows found: {len(rows)}")

    sql = """
    INSERT INTO sources (
      name, source_type, base_url, active,
      domain, product_relevance, organisation, feed_name, geo_scope, priority, meta
    )
    VALUES (
      %(name)s, %(source_type)s, %(base_url)s, true,
      %(domain)s, %(product_relevance)s, %(organisation)s, %(feed_name)s, %(geo_scope)s, %(priority)s, %(meta)s::jsonb
    )
    ON CONFLICT (source_type, base_url)
    DO UPDATE SET
      name = EXCLUDED.name,
      domain = EXCLUDED.domain,
      product_relevance = EXCLUDED.product_relevance,
      organisation = EXCLUDED.organisation,
      feed_name = EXCLUDED.feed_name,
      geo_scope = EXCLUDED.geo_scope,
      priority = EXCLUDED.priority,
      meta = sources.meta || EXCLUDED.meta;
    """

    if args.dry_run:
        for v in rows[:5]:
            print(v)
        print("Dry run only. Exiting.")
        return

    inserted = 0
    with get_conn() as conn:
        with conn.cursor() as cur:
            for v in rows:
                domain = norm(v[idx.get("Domain")]) if "Domain" in idx else None
                pr = norm(v[idx.get("Product Relevance")]) if "Product Relevance" in idx else None
                org = norm(v[idx.get("Organisation")])
                feed_name = norm(v[idx.get("Feed Name")])
                feed_type = norm(v[idx.get("Feed Type")]).lower()  # rss / atom
                feed_url = norm(v[idx.get("Feed URL")])
                geo = norm(v[idx.get("Geo Scope")]) if "Geo Scope" in idx else None
                priority = norm(v[idx.get("Priority")]) if "Priority" in idx else None

                name = f"{org} — {feed_name}"
                meta = {
                    "source_list": "master_rss_list",
                }

                cur.execute(
                    sql,
                    dict(
                        name=name,
                        source_type=feed_type,
                        base_url=feed_url,
                        domain=domain,
                        product_relevance=pr,
                        organisation=org,
                        feed_name=feed_name,
                        geo_scope=geo,
                        priority=priority,
                        meta=json.dumps(meta),
                    ),
                )
                inserted += 1
        conn.commit()

    print(f"Upserted sources: {inserted}")

if __name__ == "__main__":
    main()
